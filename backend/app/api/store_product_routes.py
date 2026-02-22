import uuid as uuid_mod
import logging
import re
from datetime import datetime
from typing import Optional, List

from fastapi import APIRouter, Depends, BackgroundTasks, Query, HTTPException, UploadFile, File
from fastapi.responses import Response
from sqlalchemy.orm import Session
from sqlalchemy import func, or_, cast, String
from pydantic import BaseModel

from app.db.database import get_db, SessionLocal
from app.db.models import StoreProduct, MonitoredProduct, ScrapeJob, ScrapeResult
from app.core.security import require_mutating_api_key

logger = logging.getLogger(__name__)

router = APIRouter(
    prefix="/api/store-products",
    tags=["Store Products"],
    dependencies=[Depends(require_mutating_api_key)],
)


@router.get("")
async def list_store_products(
    platform: Optional[str] = Query(None),
    brand: Optional[str] = Query(None),
    category: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    min_price: Optional[float] = Query(None),
    max_price: Optional[float] = Query(None),
    min_rating: Optional[float] = Query(None),
    sku: Optional[str] = Query(None),
    barcode: Optional[str] = Query(None),
    sort_by: Optional[str] = Query("created_at"),
    sort_dir: Optional[str] = Query("desc"),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    db: Session = Depends(get_db)
):
    q = db.query(StoreProduct)

    if platform:
        q = q.filter(StoreProduct.platform == platform)
    if brand:
        q = q.filter(StoreProduct.brand.ilike(f"%{brand}%"))
    if category:
        q = q.filter(StoreProduct.category.ilike(f"%{category}%"))
    if search:
        q = q.filter(or_(
            StoreProduct.product_name.ilike(f"%{search}%"),
            StoreProduct.brand.ilike(f"%{search}%"),
            StoreProduct.sku.ilike(f"%{search}%"),
        ))
    if min_price is not None:
        q = q.filter(StoreProduct.price >= min_price)
    if max_price is not None:
        q = q.filter(StoreProduct.price <= max_price)
    if min_rating is not None:
        q = q.filter(StoreProduct.rating >= min_rating)
    if sku:
        q = q.filter(StoreProduct.sku == sku)
    if barcode:
        q = q.filter(StoreProduct.barcode == barcode)

    total = q.count()

    sort_col = getattr(StoreProduct, sort_by, StoreProduct.created_at)
    if sort_dir == "asc":
        q = q.order_by(sort_col.asc())
    else:
        q = q.order_by(sort_col.desc())

    offset = (page - 1) * page_size
    products = q.offset(offset).limit(page_size).all()

    return {
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": (total + page_size - 1) // page_size if total > 0 else 0,
        "products": [_serialize_product(p) for p in products],
    }


@router.get("/filters")
async def get_filter_options(
    platform: Optional[str] = Query(None),
    db: Session = Depends(get_db)
):
    q = db.query(StoreProduct)
    if platform:
        q = q.filter(StoreProduct.platform == platform)

    brands = db.query(StoreProduct.brand, func.count(StoreProduct.id).label('cnt')).filter(
        StoreProduct.brand.isnot(None)
    )
    if platform:
        brands = brands.filter(StoreProduct.platform == platform)
    brands = brands.group_by(StoreProduct.brand).order_by(func.count(StoreProduct.id).desc()).limit(100).all()

    categories = db.query(StoreProduct.category, func.count(StoreProduct.id).label('cnt')).filter(
        StoreProduct.category.isnot(None)
    )
    if platform:
        categories = categories.filter(StoreProduct.platform == platform)
    categories = categories.group_by(StoreProduct.category).order_by(func.count(StoreProduct.id).desc()).limit(100).all()

    platforms = db.query(StoreProduct.platform, func.count(StoreProduct.id)).group_by(
        StoreProduct.platform
    ).all()

    price_stats = db.query(
        func.min(StoreProduct.price),
        func.max(StoreProduct.price),
        func.avg(StoreProduct.price),
    )
    if platform:
        price_stats = price_stats.filter(StoreProduct.platform == platform)
    price_stats = price_stats.first()

    return {
        "brands": [{"name": b[0], "count": b[1]} for b in brands],
        "categories": [{"name": c[0], "count": c[1]} for c in categories],
        "platforms": [{"name": p[0], "count": p[1]} for p in platforms],
        "price_range": {
            "min": float(price_stats[0]) if price_stats[0] else 0,
            "max": float(price_stats[1]) if price_stats[1] else 0,
            "avg": float(price_stats[2]) if price_stats[2] else 0,
        }
    }


@router.get("/stats")
async def get_store_stats(db: Session = Depends(get_db)):
    total = db.query(func.count(StoreProduct.id)).scalar()
    by_platform = db.query(
        StoreProduct.platform, func.count(StoreProduct.id)
    ).group_by(StoreProduct.platform).all()

    return {
        "total_products": total,
        "by_platform": {p[0]: p[1] for p in by_platform},
    }


@router.get("/{product_id}")
async def get_store_product(product_id: str, db: Session = Depends(get_db)):
    product = db.query(StoreProduct).filter(StoreProduct.id == uuid_mod.UUID(product_id)).first()
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    return _serialize_product(product, include_raw=True)


@router.post("/scrape-from-monitor")
async def scrape_from_price_monitor(
    background_tasks: BackgroundTasks,
    platform: Optional[str] = Query(None),
    db: Session = Depends(get_db)
):
    q = db.query(MonitoredProduct).filter(MonitoredProduct.is_active == True)
    if platform:
        q = q.filter(MonitoredProduct.platform == platform)

    products = q.all()
    if not products:
        msg = f"No active monitored products found for platform '{platform}'" if platform else "No active monitored products found"
        if platform and platform not in ('hepsiburada', 'trendyol'):
            msg += ". Only Hepsiburada and Trendyol products are available in Price Monitor. Use 'Import Excel' or 'URL Scraper' for other platforms."
        raise HTTPException(status_code=404, detail=msg)

    urls_data = []
    for p in products:
        urls_data.append({
            'url': p.product_url,
            'product_name': p.product_name,
            'barcode': p.barcode,
            'sku': p.sku,
            'platform': p.platform,
            'monitored_product_id': str(p.id),
        })

    job = ScrapeJob(total_urls=len(urls_data), status="running")
    db.add(job)
    db.commit()
    db.refresh(job)

    logger.info(f"[STORE-SCRAPE] Created job {str(job.id)[:8]} for platform={platform or 'all'}, {len(urls_data)} URLs from price monitor")

    for u in urls_data:
        result = ScrapeResult(
            scrape_job_id=job.id,
            url=u['url'],
            product_name=u.get('product_name'),
            barcode=u.get('barcode'),
            status="pending"
        )
        db.add(result)
    db.commit()

    url_metadata = {r.url: next((u for u in urls_data if u['url'] == r.url), {}) for r in db.query(ScrapeResult).filter(ScrapeResult.scrape_job_id == job.id).all()}

    logger.info(f"[STORE-SCRAPE] Job {str(job.id)[:8]} queued as background task, starting scrape...")

    background_tasks.add_task(
        run_scrape_and_save_job,
        str(job.id),
        url_metadata
    )

    return {
        "job_id": str(job.id),
        "status": "running",
        "total_urls": len(urls_data),
        "platform": platform or "all",
    }


@router.get("/scrape-job-status/{job_id}")
async def get_scrape_job_status(job_id: str, db: Session = Depends(get_db)):
    job = db.query(ScrapeJob).filter(ScrapeJob.id == uuid_mod.UUID(job_id)).first()
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")

    completed = db.query(ScrapeResult).filter(
        ScrapeResult.scrape_job_id == job.id,
        ScrapeResult.status == "completed"
    ).count()
    failed = db.query(ScrapeResult).filter(
        ScrapeResult.scrape_job_id == job.id,
        ScrapeResult.status == "failed"
    ).count()
    pending = db.query(ScrapeResult).filter(
        ScrapeResult.scrape_job_id == job.id,
        ScrapeResult.status == "pending"
    ).count()
    skipped = db.query(ScrapeResult).filter(
        ScrapeResult.scrape_job_id == job.id,
        ScrapeResult.status == "skipped"
    ).count()

    return {
        "job_id": str(job.id),
        "status": job.status,
        "total": job.total_urls or 0,
        "completed": completed,
        "failed": failed,
        "pending": pending,
        "skipped": skipped,
        "created_at": job.created_at.isoformat() if job.created_at else None,
        "completed_at": job.completed_at.isoformat() if job.completed_at else None,
    }


@router.post("/save-from-scrape-job/{job_id}")
async def save_from_scrape_job(job_id: str, db: Session = Depends(get_db)):
    job = db.query(ScrapeJob).filter(ScrapeJob.id == uuid_mod.UUID(job_id)).first()
    if not job:
        raise HTTPException(status_code=404, detail="Scrape job not found")

    results = db.query(ScrapeResult).filter(
        ScrapeResult.scrape_job_id == job.id,
        ScrapeResult.status == "completed"
    ).all()

    saved = 0
    updated = 0
    for r in results:
        if not r.scraped_data:
            continue
        result = _save_store_product_from_scrape(db, r)
        if result == 'created':
            saved += 1
        elif result == 'updated':
            updated += 1

    return {"saved": saved, "updated": updated, "total_results": len(results)}


@router.delete("/{product_id}")
async def delete_store_product(product_id: str, db: Session = Depends(get_db)):
    product = db.query(StoreProduct).filter(StoreProduct.id == uuid_mod.UUID(product_id)).first()
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    db.delete(product)
    db.commit()
    return {"success": True}


@router.delete("")
async def delete_all_store_products(
    platform: Optional[str] = Query(None),
    db: Session = Depends(get_db)
):
    q = db.query(StoreProduct)
    if platform:
        q = q.filter(StoreProduct.platform == platform)
    count = q.delete()
    db.commit()
    return {"deleted": count}


@router.post("/import-excel")
async def import_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    import openpyxl
    from io import BytesIO

    if not file.filename:
        raise HTTPException(status_code=400, detail="No file provided")
    ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else ''
    if ext not in ('xlsx', 'xls'):
        raise HTTPException(status_code=400, detail="Only .xlsx and .xls files are supported")

    MAX_SIZE = 50 * 1024 * 1024
    content = await file.read()
    if len(content) > MAX_SIZE:
        raise HTTPException(status_code=400, detail=f"File too large (max {MAX_SIZE // (1024*1024)}MB)")

    try:
        wb = openpyxl.load_workbook(BytesIO(content), read_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid Excel file: {str(e)}")

    ws = wb.active or wb[wb.sheetnames[0]]

    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    header_map = {h: i for i, h in enumerate(headers) if h}

    required = ['Url']
    missing = [h for h in required if h not in header_map]
    if missing:
        wb.close()
        raise HTTPException(status_code=400, detail=f"Missing required columns: {', '.join(missing)}. Found: {', '.join(h for h in headers if h)}")

    def cell(row, col_name):
        idx = header_map.get(col_name)
        if idx is None:
            return None
        val = ws.cell(row, idx + 1).value
        return str(val).strip() if val is not None else None

    def parse_price(val):
        if not val:
            return None
        val = str(val).strip().replace(' ', '')
        val = val.replace('.', '').replace(',', '.')
        try:
            p = float(val)
            return p if p > 0 else None
        except (ValueError, TypeError):
            return None

    created = 0
    updated = 0
    skipped = 0

    for row_num in range(2, ws.max_row + 1):
        url = cell(row_num, 'Url')
        if not url:
            skipped += 1
            continue

        sku = cell(row_num, 'StokKodu')
        barcode = cell(row_num, 'Barkodu')
        name = cell(row_num, 'Baslik')
        brand = cell(row_num, 'Marka')
        category = cell(row_num, 'Kategori')
        price = parse_price(cell(row_num, 'Fiyat1'))
        image = cell(row_num, 'Resim')
        currency_raw = cell(row_num, 'Para_Birimi')
        currency = 'TRY' if currency_raw in ('TL', 'TRY', None) else currency_raw

        hb_sku = cell(row_num, 'Hepsiburada_SKU')
        hb_price = parse_price(cell(row_num, 'HepsiBurada_Fiyat'))
        n11_price = parse_price(cell(row_num, 'N11_Fiyat'))

        cat_parts = category.replace('>', ' > ').split(' > ') if category else []
        cat_parts = [c.strip() for c in cat_parts if c.strip()]
        cat_breadcrumbs = [{'name': c, 'url': None, 'position': i + 1} for i, c in enumerate(cat_parts)] if cat_parts else None
        cat_display = ' > '.join(cat_parts) if cat_parts else None

        specs = {}
        for i in range(1, 6):
            key = cell(row_num, f'EkDetay_Tanim{i}')
            val = cell(row_num, f'EkDetay{i}')
            if key and val:
                specs[key] = val

        platform = _detect_platform(url)

        existing = None
        if sku:
            existing = db.query(StoreProduct).filter(
                StoreProduct.platform == platform,
                StoreProduct.sku == sku
            ).first()
        if not existing:
            existing = db.query(StoreProduct).filter(
                StoreProduct.source_url == url
            ).first()

        data = {
            'platform': platform,
            'source_url': url,
            'sku': sku,
            'barcode': barcode,
            'product_name': name,
            'brand': brand,
            'category': cat_display,
            'category_breadcrumbs': cat_breadcrumbs,
            'price': price,
            'currency': currency,
            'image_url': image,
            'product_specs': specs if specs else None,
            'additional_properties': {
                'hepsiburada_sku': hb_sku,
                'hepsiburada_price': hb_price,
                'n11_price': n11_price,
            } if any([hb_sku, hb_price, n11_price]) else None,
            'updated_at': datetime.utcnow(),
        }

        if existing:
            for key, val in data.items():
                if val is not None:
                    setattr(existing, key, val)
            updated += 1
        else:
            product = StoreProduct(**data)
            db.add(product)
            created += 1

        if (row_num - 1) % 100 == 0:
            db.commit()

    db.commit()
    wb.close()

    logger.info(f"[EXCEL-IMPORT] Imported {created} new, {updated} updated, {skipped} skipped from {file.filename}")

    return {
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "total_rows": ws.max_row - 1,
        "filename": file.filename,
    }


def _parse_price_str(val) -> Optional[float]:
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val) if float(val) > 0 else None
    s = str(val).strip()
    if not s:
        return None
    s = re.sub(r'[^\d.,]', '', s)
    if not s:
        return None
    if ',' in s and '.' in s:
        if s.rindex(',') > s.rindex('.'):
            s = s.replace('.', '').replace(',', '.')
        else:
            s = s.replace(',', '')
    elif ',' in s:
        parts = s.split(',')
        if len(parts) == 2 and len(parts[1]) <= 2:
            s = s.replace(',', '.')
        else:
            s = s.replace(',', '')
    try:
        p = float(s)
        return p if p > 0 else None
    except (ValueError, TypeError):
        return None


def _extract_best_price(scraped: dict) -> Optional[float]:
    price = _parse_price_str(scraped.get('price'))
    if price:
        return price

    json_ld_list = scraped.get('json_ld', [])
    for ld in json_ld_list:
        items = []
        if isinstance(ld, dict):
            if '@graph' in ld:
                items = ld['@graph'] if isinstance(ld['@graph'], list) else [ld['@graph']]
            else:
                items = [ld]
        elif isinstance(ld, list):
            items = ld
        for item in items:
            if not isinstance(item, dict):
                continue
            ld_type = item.get('@type', '')
            is_product = ld_type == 'Product' or (isinstance(ld_type, list) and 'Product' in ld_type)
            if not is_product:
                continue
            offers = item.get('offers', {})
            offer_list = [offers] if isinstance(offers, dict) else offers if isinstance(offers, list) else []
            for offer in offer_list:
                if not isinstance(offer, dict):
                    continue
                for field in ('price', 'lowPrice', 'highPrice'):
                    p = _parse_price_str(offer.get(field))
                    if p:
                        return p
                sub_offers = offer.get('offers', [])
                if isinstance(sub_offers, list):
                    for sub in sub_offers:
                        if isinstance(sub, dict):
                            p = _parse_price_str(sub.get('price'))
                            if p:
                                return p

    og = scraped.get('og_data', {})
    if isinstance(og, dict):
        p = _parse_price_str(og.get('price:amount'))
        if p:
            return p

    for key in ('original_price', 'list_price', 'sale_price'):
        p = _parse_price_str(scraped.get(key))
        if p:
            return p

    return None


def _detect_platform(url: str) -> str:
    if 'hepsiburada.com' in url:
        return 'hepsiburada'
    elif 'trendyol.com' in url:
        return 'trendyol'
    return 'web'


def _extract_image_url(scraped: dict) -> Optional[str]:
    img = scraped.get('product_image')
    if img:
        if isinstance(img, dict):
            urls = img.get('contentUrl')
            if isinstance(urls, list) and urls:
                return urls[0]
            return img.get('url') or img.get('contentUrl')
        if isinstance(img, str):
            return img
    images = scraped.get('images', [])
    if images:
        return images[0]
    og = scraped.get('og_data', {})
    return og.get('image')


def _save_store_product_from_scrape(db: Session, scrape_result: ScrapeResult) -> str:
    scraped = scrape_result.scraped_data
    if not scraped:
        return 'skipped'

    source_url = scraped.get('source_url', scrape_result.url)
    platform = _detect_platform(source_url)
    sku = scraped.get('product_sku')

    existing = None
    if sku and platform:
        existing = db.query(StoreProduct).filter(
            StoreProduct.platform == platform,
            StoreProduct.sku == sku
        ).first()
    if not existing and source_url:
        existing = db.query(StoreProduct).filter(
            StoreProduct.source_url == source_url
        ).first()

    price_val = _extract_best_price(scraped)

    rating_val = scraped.get('rating')
    try:
        rating_val = float(rating_val) if rating_val else None
    except (ValueError, TypeError):
        rating_val = None

    data = {
        'platform': platform,
        'source_url': source_url,
        'sku': sku,
        'barcode': scraped.get('product_barcode') or scrape_result.barcode,
        'product_name': scraped.get('product_name') or scrape_result.product_name,
        'brand': scraped.get('product_brand'),
        'category': scraped.get('product_category'),
        'category_breadcrumbs': scraped.get('category_breadcrumbs'),
        'price': price_val,
        'currency': scraped.get('currency'),
        'availability': scraped.get('availability'),
        'rating': rating_val,
        'rating_count': scraped.get('rating_count'),
        'review_count': scraped.get('review_count'),
        'reviews': scraped.get('reviews'),
        'image_url': _extract_image_url(scraped),
        'images': scraped.get('images'),
        'description': scraped.get('product_description'),
        'seller_name': scraped.get('seller_name'),
        'shipping_info': scraped.get('shipping_info'),
        'return_policy': scraped.get('return_policy'),
        'product_specs': scraped.get('product_specs'),
        'additional_properties': scraped.get('additional_properties'),
        'related_products': scraped.get('related_products'),
        'og_data': scraped.get('og_data'),
        'scrape_result_id': scrape_result.id,
        'raw_scraped_data': scraped,
        'updated_at': datetime.utcnow(),
    }

    if existing:
        for key, val in data.items():
            if val is not None:
                setattr(existing, key, val)
        db.commit()
        return 'updated'
    else:
        product = StoreProduct(**data)
        db.add(product)
        db.commit()
        return 'created'


async def run_scrape_and_save_job(job_id: str, url_metadata: dict):
    import time as _time
    from app.api.url_scraper_routes import run_scrape_job

    jid = job_id[:8]
    overall_start = _time.time()

    logger.info(f"[STORE-SCRAPE] [{jid}] === STARTING SCRAPE JOB === URLs: {len(url_metadata)}")

    try:
        await run_scrape_job(job_id)
    except Exception as e:
        logger.error(f"[STORE-SCRAPE] [{jid}] Scrape job failed with exception: {e}", exc_info=True)
        return

    scrape_elapsed = round(_time.time() - overall_start, 1)
    logger.info(f"[STORE-SCRAPE] [{jid}] Scrape phase completed in {scrape_elapsed}s, now saving to StoreProduct...")

    db = SessionLocal()
    try:
        job = db.query(ScrapeJob).filter(ScrapeJob.id == uuid_mod.UUID(job_id)).first()
        if not job:
            logger.error(f"[STORE-SCRAPE] [{jid}] Job not found in DB after scraping")
            return

        results = db.query(ScrapeResult).filter(
            ScrapeResult.scrape_job_id == job.id,
            ScrapeResult.status == "completed"
        ).all()

        total_results = len(results)
        created = 0
        updated = 0
        skipped = 0
        errors = 0

        logger.info(f"[STORE-SCRAPE] [{jid}] Processing {total_results} completed results...")

        for i, r in enumerate(results):
            if not r.scraped_data:
                skipped += 1
                continue

            meta = url_metadata.get(r.url, {})
            if meta.get('monitored_product_id'):
                if not r.scraped_data.get('_monitored_product_id'):
                    r.scraped_data = {**r.scraped_data, '_monitored_product_id': meta['monitored_product_id']}
            if meta.get('barcode') and not r.barcode:
                r.barcode = meta['barcode']

            try:
                result = _save_store_product_from_scrape(db, r)
                if result == 'created':
                    created += 1
                elif result == 'updated':
                    updated += 1
                else:
                    skipped += 1
            except Exception as e:
                errors += 1
                logger.error(f"[STORE-SCRAPE] [{jid}] Error saving product from {r.url[:60]}: {e}")
                db.rollback()

            if (i + 1) % 50 == 0:
                logger.info(f"[STORE-SCRAPE] [{jid}] Save progress: {i+1}/{total_results} (new:{created} upd:{updated} skip:{skipped} err:{errors})")

        total_elapsed = round(_time.time() - overall_start, 1)
        logger.info(
            f"[STORE-SCRAPE] [{jid}] === JOB COMPLETE === "
            f"Total time: {total_elapsed}s | "
            f"Scraped: {total_results}/{len(url_metadata)} | "
            f"Saved: new={created} updated={updated} skipped={skipped} errors={errors}"
        )
    except Exception as e:
        logger.error(f"[STORE-SCRAPE] [{jid}] Error in save phase: {e}", exc_info=True)
        db.rollback()
    finally:
        db.close()


def _serialize_product(p: StoreProduct, include_raw: bool = False) -> dict:
    result = {
        "id": str(p.id),
        "platform": p.platform,
        "source_url": p.source_url,
        "sku": p.sku,
        "barcode": p.barcode,
        "product_name": p.product_name,
        "brand": p.brand,
        "category": p.category,
        "category_breadcrumbs": p.category_breadcrumbs,
        "price": float(p.price) if p.price else None,
        "currency": p.currency,
        "availability": p.availability,
        "rating": p.rating,
        "rating_count": p.rating_count,
        "review_count": p.review_count,
        "reviews": p.reviews,
        "image_url": p.image_url,
        "images": p.images,
        "description": p.description,
        "seller_name": p.seller_name,
        "shipping_info": p.shipping_info,
        "return_policy": p.return_policy,
        "product_specs": p.product_specs,
        "additional_properties": p.additional_properties,
        "related_products": p.related_products,
        "created_at": p.created_at.isoformat() if p.created_at else None,
        "updated_at": p.updated_at.isoformat() if p.updated_at else None,
    }
    if include_raw:
        result["raw_scraped_data"] = p.raw_scraped_data
        result["og_data"] = p.og_data
    return result
