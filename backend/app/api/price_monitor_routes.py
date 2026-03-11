import asyncio
import csv
import io
from datetime import datetime, timedelta
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from sqlalchemy.orm import Session
from sqlalchemy import desc, func, or_, text
from typing import List, Optional, Dict, Any
from time import perf_counter

from app.db.database import get_db, SessionLocal
from app.db.models import MonitoredProduct, SellerSnapshot, PriceMonitorTask, User
from app.core.config import settings
from app.core.logger import api_logger as logger, log_endpoint_metric
from app.core.auth import get_current_user

from app.api._shared import (
    MonitoredProductInput,
    BulkProductsRequest,
    MonitoredProductResponse,
    SellerSnapshotResponse,
    ProductWithSellersResponse,
    FetchTaskResponse,
    _to_float,
    _calculate_price_alerts,
    _resolve_product_url,
    _is_valid_http_url,
    _require_scraper_api_or_503,
    _require_queue_or_503,
    _run_read_query_with_retry,
    _get_price_monitor_service,
    _get_trendyol_price_monitor_service,
    extract_sku_from_url,
)

router = APIRouter(dependencies=[Depends(get_current_user)])


@router.post("/price-monitor/products")
async def add_monitored_products(
    request: BulkProductsRequest,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """JSON formatında SKU listesi yükle - platform: hepsiburada veya trendyol"""
    added = 0
    updated = 0
    errors = []
    platform = request.platform.lower()

    # Plan limiti kontrolü
    from app.core.plan_limits import get_user_limits
    limits = get_user_limits(user)
    current_count = db.query(MonitoredProduct).filter(
        MonitoredProduct.user_id == user.id,
        MonitoredProduct.is_active == True,
    ).count()
    remaining = limits["max_skus"] - current_count
    if remaining <= 0:
        raise HTTPException(
            status_code=403,
            detail=f"SKU limitinize ulastiniz ({current_count}/{limits['max_skus']}). Planinizi yukseltin."
        )

    for item in request.products:
        try:
            sku = item.sku
            if sku:
                if sku.startswith('SKU: '):
                    sku = sku.replace('SKU: ', '')
            else:
                sku = extract_sku_from_url(item.productUrl, platform) if item.productUrl else None

            if not sku:
                errors.append({"url": item.productUrl, "error": "SKU bulunamadı"})
                continue

            resolved_product_url = _resolve_product_url(platform, sku, item.productUrl)

            existing = db.query(MonitoredProduct).filter(
                MonitoredProduct.user_id == user.id,
                MonitoredProduct.sku == sku,
                MonitoredProduct.platform == platform,
            ).first()

            if existing:
                if resolved_product_url:
                    existing.product_url = resolved_product_url
                if item.productName:
                    existing.product_name = item.productName
                if item.barcode:
                    existing.barcode = item.barcode
                if item.brand:
                    existing.brand = item.brand
                if item.price is not None:
                    existing.threshold_price = item.price
                    if item.campaignPrice is not None:
                        existing.alert_campaign_price = item.campaignPrice
                    else:
                        existing.alert_campaign_price = round(item.price * 0.9, 2)
                elif item.campaignPrice is not None:
                    existing.alert_campaign_price = item.campaignPrice
                if item.sellerStockCode:
                    existing.seller_stock_code = item.sellerStockCode
                existing.is_active = True
                updated += 1
            else:
                campaign_price = item.campaignPrice if item.campaignPrice is not None else (
                    round(item.price * 0.9, 2) if item.price else None
                )
                product = MonitoredProduct(
                    user_id=user.id,
                    platform=platform,
                    sku=sku,
                    barcode=item.barcode,
                    product_url=resolved_product_url,
                    product_name=item.productName,
                    brand=item.brand,
                    threshold_price=item.price,
                    alert_campaign_price=campaign_price,
                    seller_stock_code=item.sellerStockCode,
                    is_active=True,
                )
                db.add(product)
                added += 1
        except Exception as e:
            logger.warning(f"Import error for SKU {item.sku}: {e}")
            logger.error(f"Bulk import item failed (sku={item.sku or item.productUrl}): {type(e).__name__}: {e}")
            errors.append({"sku": item.sku or item.productUrl, "error": "Urun eklenirken hata olustu"})

    db.commit()

    return {
        "added": added,
        "updated": updated,
        "errors": errors,
        "total": len(request.products),
        "platform": platform
    }


@router.post("/price-monitor/products/import")
async def import_products_csv(
    file: UploadFile = File(...),
    platform: str = Query("hepsiburada", description="Platform: hepsiburada veya trendyol"),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """CSV veya Excel dosyasından ürün import et."""
    if not file.filename:
        raise HTTPException(status_code=400, detail="Dosya adı gerekli")

    ext = file.filename.rsplit(".", 1)[-1].lower() if "." in file.filename else ""
    if ext not in ("csv", "xlsx", "xls"):
        raise HTTPException(status_code=400, detail="Desteklenen formatlar: CSV, XLSX")

    content = await file.read()

    # Parse rows based on file type
    rows = []
    if ext == "csv":
        text_content = content.decode("utf-8-sig")
        # Auto-detect delimiter
        first_lines = text_content.split("\n", 3)[:3]
        sample = "\n".join(first_lines)
        delimiters = {";": sample.count(";"), ",": sample.count(","), "\t": sample.count("\t")}
        delimiter = max(delimiters, key=delimiters.get) if max(delimiters.values()) > 0 else ","

        reader = csv.DictReader(io.StringIO(text_content), delimiter=delimiter)
        for row in reader:
            rows.append({k.strip().lower(): v.strip() if v else "" for k, v in row.items()})
    else:
        # Excel
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
        ws = wb.active
        headers = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                headers = [str(h).strip().lower() if h else f"col_{j}" for j, h in enumerate(row)]
                continue
            row_dict = {}
            for j, val in enumerate(row):
                if j < len(headers):
                    row_dict[headers[j]] = str(val).strip() if val is not None else ""
            rows.append(row_dict)
        wb.close()

    if not rows:
        raise HTTPException(status_code=400, detail="Dosyada veri bulunamadı")

    # Column mapping — flexible header detection
    def _find_col(row: dict, candidates: list) -> str:
        for key in row:
            for candidate in candidates:
                if candidate in key:
                    return row[key]
        return ""

    added = 0
    updated = 0
    errors = []
    skipped = 0

    for i, row in enumerate(rows, start=2):  # Row 2 = first data row
        sku = _find_col(row, ["sku", "ürün kodu", "urun kodu", "product_code", "product code"])
        barcode = _find_col(row, ["barcode", "barkod"])
        name = _find_col(row, ["name", "ürün adı", "urun adi", "product_name", "product name", "ürün"])
        brand = _find_col(row, ["brand", "marka"])
        threshold = _find_col(row, ["threshold", "eşik", "esik", "threshold_price"])
        unit_cost = _find_col(row, ["cost", "maliyet", "unit_cost", "birim maliyet"])
        shipping = _find_col(row, ["shipping", "kargo", "shipping_cost", "kargo bedeli"])
        url = _find_col(row, ["url", "link", "product_url"])
        seller_stock = _find_col(row, ["seller_stock", "stok kodu", "stok_kodu", "seller_stock_code"])

        if not sku and not barcode and not url:
            skipped += 1
            continue

        # SKU yoksa URL'den çıkar
        if not sku and url:
            try:
                sku = extract_sku_from_url(url)
            except Exception:
                pass

        if not sku:
            errors.append({"row": i, "error": "SKU bulunamadı"})
            continue

        # URL yoksa oluştur
        if not url:
            try:
                url = _resolve_product_url(platform, sku)
            except Exception:
                url = (
                    f"https://www.hepsiburada.com/-p-{sku}"
                    if platform == "hepsiburada"
                    else f"https://www.trendyol.com/-p-{sku}"
                )

        # Mevcut ürün kontrolü
        existing = db.query(MonitoredProduct).filter(
            MonitoredProduct.user_id == user.id,
            MonitoredProduct.platform == platform,
            MonitoredProduct.sku == sku,
        ).first()

        if existing:
            if name and not existing.product_name:
                existing.product_name = name
            if brand and not existing.brand:
                existing.brand = brand
            if barcode and not existing.barcode:
                existing.barcode = barcode
            if threshold:
                try:
                    existing.threshold_price = float(threshold.replace(",", "."))
                except ValueError:
                    pass
            if unit_cost:
                try:
                    existing.unit_cost = float(unit_cost.replace(",", "."))
                except ValueError:
                    pass
            if shipping:
                try:
                    existing.shipping_cost = float(shipping.replace(",", "."))
                except ValueError:
                    pass
            if seller_stock:
                existing.seller_stock_code = seller_stock
            existing.is_active = True
            updated += 1
        else:
            product = MonitoredProduct(
                user_id=user.id,
                platform=platform,
                sku=sku,
                barcode=barcode or None,
                product_url=url,
                product_name=name or None,
                brand=brand or None,
                seller_stock_code=seller_stock or None,
                is_active=True,
            )
            if threshold:
                try:
                    product.threshold_price = float(threshold.replace(",", "."))
                except ValueError:
                    pass
            if unit_cost:
                try:
                    product.unit_cost = float(unit_cost.replace(",", "."))
                except ValueError:
                    pass
            if shipping:
                try:
                    product.shipping_cost = float(shipping.replace(",", "."))
                except ValueError:
                    pass
            db.add(product)
            added += 1

    try:
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Veritabanı hatası: {str(e)}")

    return {
        "success": True,
        "added": added,
        "updated": updated,
        "skipped": skipped,
        "errors": errors[:20],  # İlk 20 hata
        "total_rows": len(rows),
    }


@router.get("/price-monitor/products/import/template")
async def download_import_template():
    """Örnek CSV template indir."""
    from fastapi.responses import Response

    csv_content = "SKU,Barcode,Product Name,Brand,Threshold Price,Unit Cost,Shipping Cost,Seller Stock Code\n"
    csv_content += "HBV00001ABC12,8680001234567,Örnek Ürün Adı,Marka,99.90,50.00,15.00,STK-001\n"
    csv_content += "HBV00002DEF34,8680009876543,İkinci Ürün,Diğer Marka,149.90,80.00,15.00,STK-002\n"

    return Response(
        content=csv_content,
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=marketpulse_import_template.csv"},
    )


@router.get("/price-monitor/products")
async def get_monitored_products(
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
    active_only: bool = False,
    platform: Optional[str] = None,
    brand: Optional[str] = None,
    price_alert_only: bool = False,
    campaign_alert_only: bool = False,
    search: Optional[str] = None,
    limit: int = Query(100, ge=1, le=500),
    offset: int = Query(0, ge=0),
):
    """İzlenen ürün listesini getir - platform, marka, price/campaign alert ve arama filtresi ile"""
    start_time = perf_counter()
    # --- 1. Build product query with all SQL-level filters ---
    base_query = db.query(MonitoredProduct).filter(MonitoredProduct.user_id == user.id)
    if platform:
        base_query = base_query.filter(MonitoredProduct.platform == platform.lower())
    if brand:
        base_query = base_query.filter(MonitoredProduct.brand == brand)
    if search:
        search_lower = f"%{search.lower()}%"
        base_query = base_query.filter(
            or_(
                MonitoredProduct.sku.ilike(search_lower),
                MonitoredProduct.barcode.ilike(search_lower),
                MonitoredProduct.product_name.ilike(search_lower),
                MonitoredProduct.seller_stock_code.ilike(search_lower),
            )
        )

    query = base_query
    if active_only:
        query = query.filter(MonitoredProduct.is_active == True)

    query = query.order_by(desc(MonitoredProduct.created_at))

    # When alert filters are active we need to check all products (can't paginate before filtering)
    needs_alert_filter = price_alert_only or campaign_alert_only

    if not needs_alert_filter:
        # --- Single round-trip: products + counts + snapshots via CTE + LATERAL JOIN ---
        combined_rows = _run_read_query_with_retry(
            db,
            lambda: db.execute(
                text("""
                    WITH paged AS (
                        SELECT id, platform, sku, barcode, product_url, product_name, brand,
                               seller_stock_code, threshold_price, alert_campaign_price,
                               image_url, is_active, last_fetched_at,
                               COUNT(*) OVER() AS _total,
                               SUM(CASE WHEN is_active THEN 1 ELSE 0 END) OVER() AS _active
                        FROM monitored_products
                        WHERE (user_id = :user_id OR user_id IS NULL)
                          {platform_clause}
                          {brand_clause}
                          {active_clause}
                          {search_clause}
                        ORDER BY created_at DESC
                        LIMIT :limit OFFSET :offset
                    )
                    SELECT p.id, p.platform, p.sku, p.barcode, p.product_url, p.product_name,
                           p.brand, p.seller_stock_code, p.threshold_price, p.alert_campaign_price,
                           p.image_url, p.is_active, p.last_fetched_at, p._total, p._active,
                           s.merchant_id, s.price AS s_price, s.original_price AS s_original, s.campaign_price AS s_campaign
                    FROM paged p
                    LEFT JOIN LATERAL (
                        SELECT DISTINCT ON (merchant_id) merchant_id, price, original_price, campaign_price
                        FROM seller_snapshots ss
                        WHERE ss.monitored_product_id = p.id
                        ORDER BY merchant_id, snapshot_date DESC
                    ) s ON true
                """.format(
                    platform_clause="AND platform = :platform" if platform else "",
                    brand_clause="AND brand = :brand" if brand else "",
                    active_clause="AND is_active = true" if active_only else "",
                    search_clause="AND (sku ILIKE :search OR barcode ILIKE :search OR product_name ILIKE :search OR seller_stock_code ILIKE :search)" if search else "",
                )),
                {
                    "user_id": str(user.id),
                    **({"platform": platform.lower()} if platform else {}),
                    **({"brand": brand} if brand else {}),
                    **({"search": f"%{search.lower()}%"} if search else {}),
                    "limit": limit,
                    "offset": offset,
                },
            ).fetchall(),
            "price-monitor/products",
        )

        if not combined_rows:
            response = {"products": [], "total": 0, "active_count": 0, "inactive_count": 0, "limit": limit, "offset": offset}
            log_endpoint_metric(
                logger,
                "price-monitor/products",
                latency_ms=(perf_counter() - start_time) * 1000,
                platform=platform or "all",
                total=0,
                limit=limit,
                offset=offset,
            )
            return response

        total_count = combined_rows[0][13]
        active_count = combined_rows[0][14] or 0
        inactive_count = total_count - active_count

        # Group rows by product (each product may have multiple seller rows from LEFT JOIN)
        from collections import OrderedDict
        product_map: Dict[str, dict] = OrderedDict()
        for r in combined_rows:
            pid = str(r[0])
            if pid not in product_map:
                product_map[pid] = {
                    "id": pid, "platform": r[1], "sku": r[2], "barcode": r[3],
                    "product_url": r[4], "product_name": r[5], "brand": r[6],
                    "seller_stock_code": r[7],
                    "threshold_price": float(r[8]) if r[8] else None,
                    "alert_campaign_price": float(r[9]) if r[9] else None,
                    "image_url": r[10], "is_active": r[11],
                    "last_fetched_at": r[12].isoformat() if r[12] else None,
                    "_snapshots": [],
                }
            if r[15] is not None:  # merchant_id from LEFT JOIN
                product_map[pid]["_snapshots"].append({
                    "price": r[16], "original_price": r[17], "campaign_price": r[18],
                })

        result = []
        for p in product_map.values():
            snaps = p.pop("_snapshots")
            p["seller_count"] = len(snaps)
            threshold = p["threshold_price"]
            campaign_threshold = p["alert_campaign_price"]
            price_alert_count = 0
            campaign_alert_count = 0
            for snap in snaps:
                s_price = _to_float(snap["price"])
                s_original = _to_float(snap["original_price"])
                s_campaign = _to_float(snap["campaign_price"])
                list_price = s_original if s_original is not None else s_price
                if p["platform"] == "trendyol":
                    if threshold is not None and list_price is not None and list_price < threshold:
                        price_alert_count += 1
                    if campaign_threshold is not None and s_original is not None and s_price is not None and s_price < campaign_threshold:
                        campaign_alert_count += 1
                else:
                    if threshold is not None and list_price is not None and list_price < threshold:
                        price_alert_count += 1
                    if campaign_threshold is not None and s_campaign is not None and s_campaign < campaign_threshold:
                        campaign_alert_count += 1
            p["has_price_alert"] = price_alert_count > 0
            p["price_alert_count"] = price_alert_count
            p["has_campaign_alert"] = campaign_alert_count > 0
            p["campaign_alert_count"] = campaign_alert_count
            result.append(p)
    else:
        # Alert filter path: need all products (can't paginate before checking alerts)
        products = _run_read_query_with_retry(db, query.all, "price-monitor/products")
        if not products:
            response = {"products": [], "total": 0, "active_count": 0, "inactive_count": 0, "limit": limit, "offset": offset}
            log_endpoint_metric(
                logger,
                "price-monitor/products",
                latency_ms=(perf_counter() - start_time) * 1000,
                platform=platform or "all",
                total=0,
                limit=limit,
                offset=offset,
            )
            return response
        product_ids = [p.id for p in products]
        snap_rows = _run_read_query_with_retry(
            db,
            lambda: db.execute(
                text("""
                    SELECT DISTINCT ON (monitored_product_id, merchant_id)
                           monitored_product_id, merchant_id, price, original_price, campaign_price
                    FROM seller_snapshots
                    WHERE monitored_product_id = ANY(:product_ids)
                    ORDER BY monitored_product_id, merchant_id, snapshot_date DESC
                """),
                {"product_ids": product_ids},
            ).fetchall(),
            "price-monitor/products",
        )
        snapshot_map: Dict[str, list] = {}
        for row in snap_rows:
            pid = str(row[0])
            snapshot_map.setdefault(pid, []).append({
                "price": row[2], "original_price": row[3], "campaign_price": row[4],
            })
        result = []
        for product in products:
            pid = str(product.id)
            snaps = snapshot_map.get(pid, [])
            threshold = float(product.threshold_price) if product.threshold_price else None
            campaign_threshold = float(product.alert_campaign_price) if product.alert_campaign_price else None
            price_alert_count = 0
            campaign_alert_count = 0
            for snap in snaps:
                s_price = _to_float(snap["price"])
                s_original = _to_float(snap["original_price"])
                s_campaign = _to_float(snap["campaign_price"])
                list_price = s_original if s_original is not None else s_price
                if product.platform == "trendyol":
                    if threshold is not None and list_price is not None and list_price < threshold:
                        price_alert_count += 1
                    if campaign_threshold is not None and s_original is not None and s_price is not None and s_price < campaign_threshold:
                        campaign_alert_count += 1
                else:
                    if threshold is not None and list_price is not None and list_price < threshold:
                        price_alert_count += 1
                    if campaign_threshold is not None and s_campaign is not None and s_campaign < campaign_threshold:
                        campaign_alert_count += 1
            has_price_alert = price_alert_count > 0
            has_campaign_alert = campaign_alert_count > 0
            if price_alert_only and not has_price_alert:
                continue
            if campaign_alert_only and not has_campaign_alert:
                continue
            result.append({
                "id": pid, "platform": product.platform, "sku": product.sku,
                "barcode": product.barcode, "product_url": product.product_url,
                "product_name": product.product_name, "brand": product.brand,
                "seller_stock_code": product.seller_stock_code,
                "threshold_price": float(product.threshold_price) if product.threshold_price else None,
                "alert_campaign_price": float(product.alert_campaign_price) if product.alert_campaign_price else None,
                "image_url": product.image_url, "is_active": product.is_active,
                "last_fetched_at": product.last_fetched_at.isoformat() if product.last_fetched_at else None,
                "seller_count": len(snaps),
                "has_price_alert": has_price_alert, "price_alert_count": price_alert_count,
                "has_campaign_alert": has_campaign_alert, "campaign_alert_count": campaign_alert_count,
            })
        total_count = len(result)
        active_count = sum(1 for item in result if item["is_active"])
        inactive_count = total_count - active_count
        result = result[offset:offset + limit]

    response = {
        "products": result,
        "total": total_count,
        "active_count": active_count,
        "inactive_count": inactive_count,
        "limit": limit,
        "offset": offset
    }
    log_endpoint_metric(
        logger,
        "price-monitor/products",
        latency_ms=(perf_counter() - start_time) * 1000,
        platform=platform or "all",
        total=total_count,
        limit=limit,
        offset=offset,
    )
    return response


@router.get("/price-monitor/export")
async def export_price_monitor_data(
    platform: str = Query(..., description="Platform: hepsiburada veya trendyol"),
    active_filter: str = Query("all", description="Filtre: all, active, inactive"),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """SKU bazlı gruplandırılmış JSON olarak indir"""
    from fastapi.responses import StreamingResponse
    import json

    query = db.query(MonitoredProduct).filter(
        MonitoredProduct.user_id == user.id,
        MonitoredProduct.platform == platform.lower(),
    )

    if active_filter == "active":
        query = query.filter(MonitoredProduct.is_active == True)
    elif active_filter == "inactive":
        query = query.filter(MonitoredProduct.is_active == False)

    products = query.all()

    export_data = []

    for product in products:
        snapshots = db.query(SellerSnapshot).filter(
            SellerSnapshot.monitored_product_id == product.id
        ).order_by(desc(SellerSnapshot.snapshot_date)).all()

        sellers = []
        min_price_seller = None
        min_price_value = None
        seen_merchants = set()

        for s in snapshots:
            if s.merchant_id not in seen_merchants:
                seen_merchants.add(s.merchant_id)

                if platform.lower() == "hepsiburada":
                    merchant_url = f"https://www.hepsiburada.com/magaza/{s.merchant_url_postfix}" if s.merchant_url_postfix else ""
                elif platform.lower() == "trendyol":
                    merchant_url = f"https://www.trendyol.com{s.merchant_url_postfix}" if s.merchant_url_postfix else ""
                else:
                    merchant_url = ""

                seller_data = {
                    "merchant_name": s.merchant_name,
                    "merchant_id": s.merchant_id,
                    "merchant_url": merchant_url,
                    "merchant_rating": float(s.merchant_rating) if s.merchant_rating else None,
                    "merchant_city": s.merchant_city or "",
                    "price": float(s.price) if s.price else None,
                    "original_price": float(s.original_price) if s.original_price else None,
                    "minimum_price": float(s.minimum_price) if s.minimum_price else None,
                    "discount_rate": float(s.discount_rate) if s.discount_rate else None,
                    "stock_quantity": s.stock_quantity,
                    "buybox_order": s.buybox_order,
                    "free_shipping": s.free_shipping,
                    "fast_shipping": s.fast_shipping,
                    "delivery_info": s.delivery_info or "",
                    "campaign_info": s.campaign_info or "",
                    "snapshot_date": s.snapshot_date.isoformat() if s.snapshot_date else ""
                }
                sellers.append(seller_data)

                seller_price = seller_data["price"]
                if seller_price is not None and (min_price_value is None or seller_price < min_price_value):
                    min_price_value = seller_price
                    min_price_seller = {
                        "merchant_name": s.merchant_name,
                        "merchant_url": merchant_url,
                        "price": seller_price,
                        "buybox_order": s.buybox_order
                    }

        product_data = {
            "platform": product.platform,
            "sku": product.sku,
            "barcode": product.barcode or "",
            "product_name": product.product_name or "",
            "product_url": product.product_url or "",
            "is_active": product.is_active,
            "min_price": min_price_seller["price"] if min_price_seller else None,
            "min_price_seller": min_price_seller,
            "seller_count": len(sellers),
            "sellers": sellers
        }
        export_data.append(product_data)

    output = json.dumps(export_data, ensure_ascii=False, indent=2)
    filename = f"price_monitor_{platform}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.json"

    return StreamingResponse(
            iter([output]),
            media_type="application/json",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )


@router.get("/price-monitor/export-excel")
async def export_price_monitor_excel(
    platform: str = Query(None, description="Platform filtresi (opsiyonel): hepsiburada veya trendyol"),
    active_filter: str = Query("all", description="Filtre: all, active, inactive"),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Ürünleri pazaryeri kategori ağacı ile Excel dosyasına aktar.

    Her platform ayrı sheet olarak eklenir.  Kategori bilgisi sırasıyla
    store_products, products ve monitored_products.product_url'den çıkarılır.
    """
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from collections import defaultdict

    platforms = [platform.lower()] if platform else ["hepsiburada", "trendyol"]

    wb = Workbook()
    wb.remove(wb.active)  # Boş varsayılan sheet'i kaldır

    # ---------- Kategori çözümleme yardımcıları ----------
    def _category_from_store_products(db: Session, user_id, mp: MonitoredProduct) -> Optional[str]:
        """store_products tablosundan kategori breadcrumbs bul."""
        from app.db.models import StoreProduct
        filters = [StoreProduct.user_id == user_id, StoreProduct.platform == mp.platform]
        if mp.barcode:
            sp = db.query(StoreProduct).filter(*filters, StoreProduct.barcode == mp.barcode).first()
            if sp:
                if sp.category_breadcrumbs:
                    names = [c.get("name") or c.get("text", "") for c in sp.category_breadcrumbs if isinstance(c, dict)]
                    if names:
                        return " > ".join(n for n in names if n)
                if sp.category:
                    return sp.category
        if mp.sku:
            sp = db.query(StoreProduct).filter(*filters, StoreProduct.sku == mp.sku).first()
            if sp:
                if sp.category_breadcrumbs:
                    names = [c.get("name") or c.get("text", "") for c in sp.category_breadcrumbs if isinstance(c, dict)]
                    if names:
                        return " > ".join(n for n in names if n)
                if sp.category:
                    return sp.category
        return None

    def _category_from_products_table(db: Session, mp: MonitoredProduct) -> Optional[str]:
        """products tablosundan kategori bul (keyword search sonuçları)."""
        from app.db.models import Product
        if mp.sku:
            p = db.query(Product).filter(
                Product.platform == mp.platform,
                Product.external_id == mp.sku,
            ).first()
            if p:
                return p.category_hierarchy or p.category_path
        return None

    def _category_from_url(url: str) -> Optional[str]:
        """Trendyol URL'inden kategori çıkar (slug'dan)."""
        if not url:
            return None
        import re
        # Trendyol URL: /brand/category-slug-p-123456
        m = re.search(r'trendyol\.com/([^/]+)/([^?]+)-p-\d+', url)
        if m:
            brand_slug = m.group(1).replace('-', ' ').title()
            return brand_slug
        return None

    def _resolve_category(db: Session, user_id, mp: MonitoredProduct) -> str:
        cat = _category_from_store_products(db, user_id, mp)
        if cat:
            return cat
        cat = _category_from_products_table(db, mp)
        if cat:
            return cat
        cat = _category_from_url(mp.product_url)
        if cat:
            return cat
        return "Kategorisiz"

    # ---------- Stil tanımları ----------
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2B3544", end_color="2B3544", fill_type="solid")
    category_font = Font(bold=True, size=11, color="FF6000")
    category_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
    thin_border = Border(
        bottom=Side(style="thin", color="DDDDDD"),
    )

    summary_data: Dict[str, Dict[str, int]] = {}  # platform -> {category: count}

    for plat in platforms:
        query = db.query(MonitoredProduct).filter(
            MonitoredProduct.user_id == user.id,
            MonitoredProduct.platform == plat,
        )
        if active_filter == "active":
            query = query.filter(MonitoredProduct.is_active == True)
        elif active_filter == "inactive":
            query = query.filter(MonitoredProduct.is_active == False)

        products = query.order_by(MonitoredProduct.brand, MonitoredProduct.product_name).all()
        if not products:
            continue

        # Son snapshot fiyatlarını toplu çek
        product_ids = [p.id for p in products]
        snap_rows = db.execute(
            text("""
                SELECT DISTINCT ON (monitored_product_id, merchant_id)
                       monitored_product_id, merchant_id, merchant_name, price, original_price,
                       campaign_price, buybox_order, free_shipping, discount_rate
                FROM seller_snapshots
                WHERE monitored_product_id = ANY(:pids)
                ORDER BY monitored_product_id, merchant_id, snapshot_date DESC
            """),
            {"pids": product_ids},
        ).fetchall()

        snap_map: Dict[str, list] = defaultdict(list)
        for row in snap_rows:
            snap_map[str(row[0])].append({
                "merchant_name": row[2],
                "price": float(row[3]) if row[3] else None,
                "original_price": float(row[4]) if row[4] else None,
                "campaign_price": float(row[5]) if row[5] else None,
                "buybox_order": row[6],
                "free_shipping": row[7],
                "discount_rate": float(row[8]) if row[8] else None,
            })

        # Kategori bazlı gruplama
        cat_products: Dict[str, list] = defaultdict(list)
        for mp in products:
            cat = _resolve_category(db, user.id, mp)
            cat_products[cat].append(mp)

        summary_data[plat] = {cat: len(prods) for cat, prods in cat_products.items()}

        # Sheet oluştur
        sheet_name = "Hepsiburada" if plat == "hepsiburada" else "Trendyol"
        ws = wb.create_sheet(title=sheet_name)

        headers = [
            "Kategori", "Ürün Adı", "SKU", "Barkod", "Marka",
            "Stok Kodu", "Durum", "Satıcı Sayısı",
            "En Düşük Fiyat", "Buybox Satıcı", "Buybox Fiyat",
            "Eşik Fiyat", "Kampanya Eşik",
        ]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.freeze_panes = "A2"

        row_num = 2
        sorted_cats = sorted(cat_products.keys(), key=lambda c: (c == "Kategorisiz", c))

        for cat in sorted_cats:
            prods = cat_products[cat]

            # Kategori başlık satırı
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=len(headers))
            cat_cell = ws.cell(row=row_num, column=1, value=f"{cat}  ({len(prods)} ürün)")
            cat_cell.font = category_font
            cat_cell.fill = category_fill
            cat_cell.alignment = Alignment(horizontal="left", vertical="center")
            row_num += 1

            for mp in prods:
                snaps = snap_map.get(str(mp.id), [])
                seller_count = len(snaps)

                # En düşük fiyat
                prices = [s["price"] for s in snaps if s["price"] is not None]
                min_price = min(prices) if prices else None

                # Buybox winner
                buybox = next((s for s in snaps if s.get("buybox_order") == 0), None)
                if not buybox and snaps:
                    sorted_snaps = sorted(snaps, key=lambda s: s.get("buybox_order") or 999)
                    buybox = sorted_snaps[0]

                ws.cell(row=row_num, column=1, value=cat)
                ws.cell(row=row_num, column=2, value=mp.product_name or "")
                ws.cell(row=row_num, column=3, value=mp.sku or "")
                ws.cell(row=row_num, column=4, value=mp.barcode or "")
                ws.cell(row=row_num, column=5, value=mp.brand or "")
                ws.cell(row=row_num, column=6, value=mp.seller_stock_code or "")
                ws.cell(row=row_num, column=7, value="Aktif" if mp.is_active else "İnaktif")
                ws.cell(row=row_num, column=8, value=seller_count)
                ws.cell(row=row_num, column=9, value=min_price)
                ws.cell(row=row_num, column=10, value=buybox["merchant_name"] if buybox else "")
                ws.cell(row=row_num, column=11, value=buybox["price"] if buybox else None)
                ws.cell(row=row_num, column=12, value=float(mp.threshold_price) if mp.threshold_price else None)
                ws.cell(row=row_num, column=13, value=float(mp.alert_campaign_price) if mp.alert_campaign_price else None)

                for col_idx in range(1, len(headers) + 1):
                    ws.cell(row=row_num, column=col_idx).border = thin_border

                row_num += 1

        # Sütun genişlikleri
        col_widths = [30, 45, 18, 16, 20, 15, 10, 12, 14, 25, 14, 14, 14]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # ---------- Özet Sheet ----------
    if summary_data:
        ws_summary = wb.create_sheet(title="Kategori Özeti", index=0)

        sum_headers = ["Platform", "Kategori", "Ürün Sayısı"]
        for col_idx, h in enumerate(sum_headers, 1):
            cell = ws_summary.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        ws_summary.freeze_panes = "A2"
        row_num = 2

        for plat, cats in summary_data.items():
            plat_label = "Hepsiburada" if plat == "hepsiburada" else "Trendyol"
            for cat in sorted(cats.keys(), key=lambda c: (c == "Kategorisiz", c)):
                ws_summary.cell(row=row_num, column=1, value=plat_label)
                ws_summary.cell(row=row_num, column=2, value=cat)
                ws_summary.cell(row=row_num, column=3, value=cats[cat])
                row_num += 1

        ws_summary.column_dimensions["A"].width = 18
        ws_summary.column_dimensions["B"].width = 50
        ws_summary.column_dimensions["C"].width = 14

    if not wb.sheetnames:
        ws = wb.create_sheet(title="Boş")
        ws.cell(row=1, column=1, value="Seçilen filtrelere uygun ürün bulunamadı.")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    plat_suffix = platform or "all"
    filename = f"price_monitor_categories_{plat_suffix}_{timestamp}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/price-monitor/brands")
async def get_monitored_product_brands(
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
    platform: Optional[str] = None,
):
    """Platform için marka listesini getir"""
    from sqlalchemy import distinct
    query = db.query(distinct(MonitoredProduct.brand)).filter(
        MonitoredProduct.user_id == user.id,
        MonitoredProduct.brand.isnot(None),
    )
    if platform:
        query = query.filter(MonitoredProduct.platform == platform.lower())
    brands = [b[0] for b in query.all() if b[0]]
    return {"brands": sorted(brands)}


@router.get("/price-monitor/products/{product_id}/price-history")
async def get_price_history(
    product_id: str,
    days: int = Query(30, ge=1, le=365, description="Kaç günlük geçmiş"),
    merchant_id: Optional[str] = Query(None, description="Belirli satıcı filtresi"),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Ürünün fiyat geçmişi — satıcı bazlı trend verisi."""
    import uuid as uuid_mod

    product = db.query(MonitoredProduct).filter(
        MonitoredProduct.id == uuid_mod.UUID(product_id),
        MonitoredProduct.user_id == user.id,
    ).first()

    if not product:
        raise HTTPException(status_code=404, detail="Ürün bulunamadı")

    since = datetime.utcnow() - timedelta(days=days)

    query = db.query(SellerSnapshot).filter(
        SellerSnapshot.monitored_product_id == product.id,
        SellerSnapshot.snapshot_date >= since,
    )

    if merchant_id:
        query = query.filter(SellerSnapshot.merchant_id == merchant_id)

    snapshots = query.order_by(SellerSnapshot.snapshot_date.asc()).all()

    # Satıcı bazlı gruplama
    merchants: Dict[str, dict] = {}
    buybox_timeline = []

    for s in snapshots:
        mid = s.merchant_id
        if mid not in merchants:
            merchants[mid] = {
                "merchant_id": mid,
                "merchant_name": s.merchant_name,
                "data_points": [],
            }

        merchants[mid]["data_points"].append({
            "date": s.snapshot_date.isoformat(),
            "price": float(s.price) if s.price else None,
            "original_price": float(s.original_price) if s.original_price else None,
            "campaign_price": float(s.campaign_price) if s.campaign_price else None,
            "buybox_order": s.buybox_order,
            "stock_quantity": s.stock_quantity,
        })

        # Buybox winner tracking
        if s.buybox_order == 1:
            buybox_timeline.append({
                "date": s.snapshot_date.isoformat(),
                "merchant_id": mid,
                "merchant_name": s.merchant_name,
                "price": float(s.price) if s.price else None,
            })

    # Min/max/avg hesaplama per merchant
    for mid, mdata in merchants.items():
        prices = [dp["price"] for dp in mdata["data_points"] if dp["price"] is not None]
        if prices:
            mdata["min_price"] = min(prices)
            mdata["max_price"] = max(prices)
            mdata["avg_price"] = round(sum(prices) / len(prices), 2)
            mdata["current_price"] = prices[-1]
            mdata["price_change"] = round(prices[-1] - prices[0], 2) if len(prices) > 1 else 0
            mdata["price_change_pct"] = (
                round((prices[-1] - prices[0]) / prices[0] * 100, 2)
                if len(prices) > 1 and prices[0] > 0
                else 0
            )

    return {
        "product_id": str(product.id),
        "product_name": product.product_name,
        "sku": product.sku,
        "platform": product.platform,
        "days": days,
        "total_snapshots": len(snapshots),
        "merchants": list(merchants.values()),
        "buybox_timeline": buybox_timeline,
    }


@router.get("/price-monitor/products/{product_id}/price-summary")
async def get_price_summary(
    product_id: str,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Ürünün güncel fiyat özeti — son snapshot'tan."""
    import uuid as uuid_mod

    product = db.query(MonitoredProduct).filter(
        MonitoredProduct.id == uuid_mod.UUID(product_id),
        MonitoredProduct.user_id == user.id,
    ).first()

    if not product:
        raise HTTPException(status_code=404, detail="Ürün bulunamadı")

    # En son snapshot tarihini bul
    latest_date = db.query(func.max(SellerSnapshot.snapshot_date)).filter(
        SellerSnapshot.monitored_product_id == product.id,
    ).scalar()

    if not latest_date:
        return {"product_id": str(product.id), "sellers": [], "message": "Henüz veri yok"}

    # Son snapshot'taki satıcıları getir (30 dk tolerans)
    latest_snapshots = db.query(SellerSnapshot).filter(
        SellerSnapshot.monitored_product_id == product.id,
        SellerSnapshot.snapshot_date >= latest_date - timedelta(minutes=30),
    ).order_by(SellerSnapshot.buybox_order.asc().nullslast()).all()

    # Deduplicate — aynı merchant_id için en son snapshot'ı al
    seen: Dict[str, Any] = {}
    for s in latest_snapshots:
        if s.merchant_id not in seen or s.snapshot_date > seen[s.merchant_id].snapshot_date:
            seen[s.merchant_id] = s

    sellers = []
    for s in sorted(seen.values(), key=lambda x: x.buybox_order or 999):
        sellers.append({
            "merchant_id": s.merchant_id,
            "merchant_name": s.merchant_name,
            "price": float(s.price) if s.price else None,
            "original_price": float(s.original_price) if s.original_price else None,
            "campaign_price": float(s.campaign_price) if s.campaign_price else None,
            "buybox_order": s.buybox_order,
            "stock_quantity": s.stock_quantity,
            "free_shipping": s.free_shipping,
            "campaigns": s.campaigns,
        })

    return {
        "product_id": str(product.id),
        "product_name": product.product_name,
        "sku": product.sku,
        "last_fetched": latest_date.isoformat() if latest_date else None,
        "seller_count": len(sellers),
        "sellers": sellers,
        "buybox_winner": sellers[0]["merchant_name"] if sellers else None,
        "lowest_price": min((s["price"] for s in sellers if s["price"]), default=None),
    }


@router.get("/price-monitor/products/{product_id}")
async def get_monitored_product_detail(
    product_id: str,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Tek bir ürün ve satıcılarını getir"""
    product = db.query(MonitoredProduct).filter(
        MonitoredProduct.id == product_id,
        MonitoredProduct.user_id == user.id,
    ).first()

    if not product:
        raise HTTPException(status_code=404, detail="Ürün bulunamadı")

    threshold = float(product.threshold_price) if product.threshold_price else None
    campaign_threshold = float(product.alert_campaign_price) if product.alert_campaign_price else None

    latest_snapshots = db.query(SellerSnapshot).filter(
        SellerSnapshot.monitored_product_id == product.id
    ).order_by(desc(SellerSnapshot.snapshot_date)).all()

    seen_merchants = set()
    unique_sellers = []
    price_alert_count = 0
    campaign_alert_count = 0

    for s in latest_snapshots:
        if s.merchant_id not in seen_merchants:
            seen_merchants.add(s.merchant_id)
            if product.platform == "hepsiburada":
                merchant_url = f"https://www.hepsiburada.com/magaza/{s.merchant_url_postfix}" if s.merchant_url_postfix else None
            elif product.platform == "trendyol":
                merchant_url = f"https://www.trendyol.com{s.merchant_url_postfix}" if s.merchant_url_postfix else None
            else:
                merchant_url = None

            alerts = _calculate_price_alerts(product.platform, s, threshold, campaign_threshold)
            list_price = alerts["list_price"]
            selling_price = alerts["selling_price"]
            has_price_alert = alerts["has_price_alert"]
            has_campaign_alert = alerts["has_campaign_alert"]

            if has_price_alert:
                price_alert_count += 1
            if has_campaign_alert:
                campaign_alert_count += 1

            unique_sellers.append({
                "merchant_id": s.merchant_id,
                "merchant_name": s.merchant_name,
                "merchant_logo": s.merchant_logo,
                "merchant_url_postfix": s.merchant_url_postfix,
                "merchant_url": merchant_url,
                "merchant_rating": float(s.merchant_rating) if s.merchant_rating else None,
                "merchant_rating_count": s.merchant_rating_count,
                "merchant_city": s.merchant_city,
                "price": selling_price,  # Güncel satış fiyatı (kampanyalı olabilir)
                "list_price": list_price,  # Liste fiyatı (threshold ile karşılaştırılır)
                "original_price": alerts["original_price"],  # Raw original_price
                "minimum_price": float(s.minimum_price) if s.minimum_price else None,
                "discount_rate": s.discount_rate,
                "stock_quantity": s.stock_quantity,
                "buybox_order": s.buybox_order,
                "free_shipping": s.free_shipping,
                "fast_shipping": s.fast_shipping,
                "is_fulfilled_by_hb": s.is_fulfilled_by_hb,
                "campaigns": s.campaigns if s.campaigns else [],
                "campaign_price": alerts["campaign_price"],
                "snapshot_date": s.snapshot_date.isoformat(),
                "price_alert": has_price_alert,
                "campaign_alert": has_campaign_alert
            })

    return {
        "product": {
            "id": str(product.id),
            "platform": product.platform,
            "sku": product.sku,
            "barcode": product.barcode,
            "product_url": product.product_url,
            "product_name": product.product_name,
            "brand": product.brand,
            "seller_stock_code": product.seller_stock_code,
            "threshold_price": threshold,
            "alert_campaign_price": campaign_threshold,
            "image_url": product.image_url,
            "is_active": product.is_active,
            "last_fetched_at": product.last_fetched_at.isoformat() if product.last_fetched_at else None,
            "seller_count": len(unique_sellers),
            "has_price_alert": price_alert_count > 0,
            "price_alert_count": price_alert_count,
            "has_campaign_alert": campaign_alert_count > 0,
            "campaign_alert_count": campaign_alert_count
        },
        "sellers": unique_sellers
    }


@router.delete("/price-monitor/products/{product_id}")
async def delete_monitored_product(
    product_id: str,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """İzlenen ürünü sil"""
    product = db.query(MonitoredProduct).filter(
        MonitoredProduct.id == product_id,
        MonitoredProduct.user_id == user.id,
    ).first()

    if not product:
        raise HTTPException(status_code=404, detail="Ürün bulunamadı")

    db.delete(product)
    db.commit()

    return {"success": True, "message": "Ürün silindi"}


@router.delete("/price-monitor/products/bulk/all")
async def delete_all_monitored_products(
    platform: str = Query(..., description="Platform: hepsiburada veya trendyol"),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Belirli platformdaki tüm izlenen ürünleri sil"""
    base = db.query(MonitoredProduct).filter(
        MonitoredProduct.user_id == user.id,
        MonitoredProduct.platform == platform,
    )
    # İlişkili snapshot'ları önce sil (cascade yerine explicit)
    product_ids = [p.id for p in base.all()]
    if not product_ids:
        return {"success": True, "deleted_count": 0, "message": "Silinecek ürün bulunamadı"}

    db.query(SellerSnapshot).filter(
        SellerSnapshot.monitored_product_id.in_(product_ids)
    ).delete(synchronize_session=False)
    count = base.delete(synchronize_session=False)
    db.commit()

    return {"success": True, "deleted_count": count, "message": f"{count} ürün silindi"}


@router.delete("/price-monitor/products/bulk/inactive")
async def delete_inactive_monitored_products(
    platform: str = Query(..., description="Platform: hepsiburada veya trendyol"),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Belirli platformdaki inaktif ürünleri sil"""
    base = db.query(MonitoredProduct).filter(
        MonitoredProduct.user_id == user.id,
        MonitoredProduct.platform == platform,
        MonitoredProduct.is_active == False,
    )
    product_ids = [p.id for p in base.all()]
    if not product_ids:
        return {"success": True, "deleted_count": 0, "message": "Silinecek inaktif ürün bulunamadı"}

    db.query(SellerSnapshot).filter(
        SellerSnapshot.monitored_product_id.in_(product_ids)
    ).delete(synchronize_session=False)
    count = base.delete(synchronize_session=False)
    db.commit()

    return {"success": True, "deleted_count": count, "message": f"{count} inaktif ürün silindi"}


async def run_fetch_task(task_id: str, platform: str, product_ids: List[str] = None, fetch_type: str = "active"):
    """Arka planda satıcı fiyatlarını çek"""
    db = SessionLocal()
    try:
        task = db.query(PriceMonitorTask).filter(PriceMonitorTask.id == task_id).first()
        try:
            _require_scraper_api_or_503()
        except HTTPException as exc:
            if task:
                task.status = "failed"
                task.error_message = str(exc.detail)
                task.completed_at = datetime.utcnow()
                db.commit()
            return
        if task:
            if platform == "trendyol":
                await _get_trendyol_price_monitor_service().fetch_all_products(db, task, product_ids, platform, fetch_type)
            else:
                await _get_price_monitor_service().fetch_all_products(db, task, product_ids, platform, fetch_type)
    except Exception as exc:
        logger.error("Background fetch task failed task_id=%s: %s", task_id, exc)
        try:
            task = db.query(PriceMonitorTask).filter(PriceMonitorTask.id == task_id).first()
            if task and task.status not in ("completed", "failed", "stopped"):
                task.status = "failed"
                task.error_message = str(exc)[:500]
                task.completed_at = datetime.utcnow()
                db.commit()
        except Exception:
            pass
    finally:
        db.close()


@router.post("/price-monitor/fetch")
async def start_fetch_task(
    platform: str = Query("hepsiburada", description="Platform: hepsiburada veya trendyol"),
    fetch_type: str = Query("active", description="Fetch type: active, last_inactive, inactive"),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
    product_ids: Optional[List[str]] = None,
):
    """Belirli platform için satıcı fiyatlarını çekmeye başla

    fetch_type:
    - active: Aktif ürünleri çek (varsayılan)
    - last_inactive: Son fetch'te inactive olan ürünleri tekrar dene
    - inactive: Tüm inactive ürünleri çek
    """
    _require_scraper_api_or_503()
    executor = settings.price_monitor_executor()
    if executor != "local":
        _require_queue_or_503()

    # Aynı platform için zaten running bir task varsa engelle
    existing_running = db.query(PriceMonitorTask).filter(
        PriceMonitorTask.user_id == user.id,
        PriceMonitorTask.platform == platform,
        PriceMonitorTask.status.in_(["pending", "running"]),
    ).first()
    if existing_running:
        raise HTTPException(
            status_code=409,
            detail=f"{platform} için zaten devam eden bir fetch işlemi var (task_id: {existing_running.id})"
        )

    task = PriceMonitorTask(user_id=user.id, platform=platform, status="pending", fetch_type=fetch_type)
    db.add(task)
    db.commit()
    db.refresh(task)

    if executor == "local":
        bg = asyncio.create_task(run_fetch_task(str(task.id), platform, product_ids, fetch_type))
        bg.add_done_callback(lambda t: t.exception() if not t.cancelled() and t.exception() else None)
    else:
        try:
            from app.tasks import send_price_monitor_task
            send_price_monitor_task(str(task.id), platform, fetch_type, product_ids)
        except Exception as exc:
            task.status = "failed"
            task.error_message = f"Celery enqueue failed: {exc}"
            task.completed_at = datetime.utcnow()
            db.commit()
            raise HTTPException(status_code=503, detail="Fetch queue unavailable. Check Redis/Celery worker.")

    return {
        "task_id": str(task.id),
        "platform": platform,
        "fetch_type": fetch_type,
        "executor": executor,
        "status": "started",
        "message": f"{platform.capitalize()} için {fetch_type} ürünler çekme işlemi başlatıldı"
    }


@router.post("/price-monitor/fetch/{task_id}/stop")
async def stop_fetch_task(
    task_id: str,
    db: Session = Depends(get_db)
):
    """Fiyat çekme görevini durdur"""
    task = db.query(PriceMonitorTask).filter(PriceMonitorTask.id == task_id).first()

    if not task:
        raise HTTPException(status_code=404, detail="Görev bulunamadı")

    if task.status not in ["pending", "running"]:
        raise HTTPException(status_code=400, detail="Görev zaten tamamlanmış veya durdurulmuş")

    task.stop_requested = True
    db.commit()

    return {
        "success": True,
        "message": "Durdurma isteği gönderildi, mevcut ürün tamamlandıktan sonra duracak"
    }


@router.get("/price-monitor/fetch/active")
async def get_active_fetch_task(
    platform: str = Query("hepsiburada"),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Kullanıcının aktif (running/pending) fetch task'ını döndür."""
    task = db.query(PriceMonitorTask).filter(
        PriceMonitorTask.user_id == user.id,
        PriceMonitorTask.platform == platform,
        PriceMonitorTask.status.in_(["pending", "running"]),
    ).order_by(desc(PriceMonitorTask.created_at)).first()

    if not task:
        return {"active": False}

    return {
        "active": True,
        "id": str(task.id),
        "status": task.status,
        "total_products": task.total_products,
        "completed_products": task.completed_products,
        "failed_products": task.failed_products,
        "fetch_type": task.fetch_type,
        "created_at": task.created_at.isoformat(),
    }


@router.get("/price-monitor/fetch/{task_id}")
async def get_fetch_task_status(
    task_id: str,
    db: Session = Depends(get_db)
):
    """Fiyat çekme görevinin durumunu getir"""
    task = db.query(PriceMonitorTask).filter(PriceMonitorTask.id == task_id).first()

    if not task:
        raise HTTPException(status_code=404, detail="Görev bulunamadı")

    return {
        "id": str(task.id),
        "status": task.status,
        "total_products": task.total_products,
        "completed_products": task.completed_products,
        "failed_products": task.failed_products,
        "fetch_type": task.fetch_type,
        "executor": settings.price_monitor_executor(),
        "last_inactive_count": len(task.last_inactive_skus) if task.last_inactive_skus else 0,
        "created_at": task.created_at.isoformat(),
        "completed_at": task.completed_at.isoformat() if task.completed_at else None
    }


@router.get("/price-monitor/last-inactive")
async def get_last_inactive_skus(
    platform: str = Query("hepsiburada"),
    db: Session = Depends(get_db)
):
    """Son tamamlanan fetch görevinde inactive olan SKU'ları getir"""
    start_time = perf_counter()

    def _query_last_inactive() -> Dict[str, Any]:
        last_task = db.query(PriceMonitorTask).filter(
            PriceMonitorTask.platform == platform,
            PriceMonitorTask.status == "completed"
        ).order_by(PriceMonitorTask.completed_at.desc()).first()

        if not last_task:
            return {"skus": [], "count": 0, "task_id": None}

        skus = last_task.last_inactive_skus or []

        products = []
        if skus:
            product_records = db.query(MonitoredProduct).filter(
                MonitoredProduct.sku.in_(skus),
                MonitoredProduct.platform == platform
            ).all()

            products = [
                {
                    "id": str(p.id),
                    "sku": p.sku,
                    "product_name": p.product_name,
                    "brand": p.brand,
                    "is_active": p.is_active
                }
                for p in product_records
            ]

        return {
            "skus": skus,
            "count": len(skus),
            "products": products,
            "task_id": str(last_task.id),
            "completed_at": last_task.completed_at.isoformat() if last_task.completed_at else None
        }

    response = _run_read_query_with_retry(db, _query_last_inactive, "price-monitor/last-inactive")
    log_endpoint_metric(
        logger,
        "price-monitor/last-inactive",
        latency_ms=(perf_counter() - start_time) * 1000,
        platform=platform,
        count=response.get("count", 0),
    )
    return response


@router.post("/price-monitor/fetch-single/{product_id}")
async def fetch_single_product(
    product_id: str,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Tek bir ürün için satıcı fiyatlarını çek - platform otomatik belirlenir"""
    _require_scraper_api_or_503()

    product = db.query(MonitoredProduct).filter(
        MonitoredProduct.id == product_id,
        MonitoredProduct.user_id == user.id,
    ).first()

    if not product:
        raise HTTPException(status_code=404, detail="Ürün bulunamadı")

    if product.platform == "trendyol":
        success = await _get_trendyol_price_monitor_service().fetch_and_save_product(db, product)
    else:
        success = await _get_price_monitor_service().fetch_and_save_product(db, product)

    if success:
        return {"success": True, "message": f"{product.sku} için satıcı verileri güncellendi", "platform": product.platform}
    else:
        raise HTTPException(status_code=500, detail="Satıcı verileri çekilemedi")
